import re
from datetime import datetime, date
from pathlib import Path

import openpyxl

from .db import connect, init_db

LAPORAN_DIR = Path(__file__).resolve().parent.parent / "laporan"
PENJUALAN_DIR = Path(__file__).resolve().parent.parent / "penjualan"

# Modul D spec: Expense -> kategori makro
COST_CATEGORIES = {
    "HPP & Bahan Baku": [
        "Biaya Pemakaian Bahan Habis Pakai", "HPP Produk", "HPP Minuman",
    ],
    "Operasional & Tetap": [
        "Biaya Listrik", "Biaya RnM", "Biaya Sewa IT (Olsera)",
        "Biaya Sewa Perlatan", "Biaya Telepon  dan Internet",
        "Biaya Administrasi",
    ],
    "SDM & Pengembangan": [
        "Biaya SPV Visit", "Biaya Model Class Development", "Biaya Grading",
        "BPJS", "THR", "Biaya Battle Fade",
    ],
    "Marketing & Akuisisi": ["Biaya Iklan", "Biaya Sample"],
}

EXPENSE_ALIASES = {
    "biaya sewa peralatan": "Biaya Sewa Perlatan",
    "biaya telepon dan internet": "Biaya Telepon  dan Internet",
}


def _norm_outlet(v) -> str:
    s = str(v or "").strip().upper()
    return "BK21" if s.startswith("BK21") else s


KEYWORD_FALLBACK = [
    ("SDM & Pengembangan", ["bpjs", "thr", "gaji", "bonus", "poin", "vitamin",
                            "grading", "spv", "battle fade", "model class",
                            "regenerasi", "uph", "gathering", "quality visit",
                            "perbaikan skill"]),
    ("Operasional & Tetap", ["listrik", "air ", "pdam", "internet", "telepon",
                             "sewa", "rnm", "r&m", "administrasi", "laundry",
                             "security", "overhead"]),
    ("Marketing & Akuisisi", ["iklan", "sample", "desain", "merchandise",
                              "reklame", "marketing"]),
    ("HPP & Bahan Baku", ["habis pakai", "hpp"]),
]


def category_for(expense: str) -> str:
    e = EXPENSE_ALIASES.get(expense.strip().lower(), expense)
    low = e.lower()
    for cat, members in COST_CATEGORIES.items():
        if e in members or expense in members:
            return cat
    for cat, keywords in KEYWORD_FALLBACK:
        if any(k in low for k in keywords):
            return cat
    return "Lainnya"


def _iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def _num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def _header_map(ws, keys):
    """Find a row containing all keys; return {key: col_idx}."""
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).strip() if c else "" for c in row]
        if all(any(k == c for c in cells) for k in keys):
            return r, {k: cells.index(k) for k in keys}
    raise ValueError(f"header dengan {keys} tidak ditemukan di {ws.title}")


def ingest_file(path: Path, conn):
    source = str(path.relative_to(LAPORAN_DIR))
    outlet_m = re.match(r"\d+\.(BK\d+)_", path.name)
    file_outlet = outlet_m.group(1) if outlet_m else "UNKNOWN"
    cost_rows: list[tuple] = []
    rev_rows: list[tuple] = []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # --- Cost ---
    if "Cost" in wb.sheetnames:
        ws = wb["Cost"]
        rows = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        idx = {name: hdr.index(name) for name in
               ("No_Akun", "Tanggal", "Outlet", "Expense", "Deskripsi", "Total")}
        cost_rows = []
        max_cidx = max(idx.values())
        for i, row in enumerate(rows[1:], start=1):
            if len(row) <= max_cidx:
                continue
            tgl = _iso(row[idx["Tanggal"]])
            total = row[idx["Total"]]
            if not tgl or not isinstance(total, (int, float)):
                continue
            no_akun_v = row[idx["No_Akun"]]
            no_akun = int(no_akun_v) if isinstance(no_akun_v, (int, float)) else None
            outlet = row[idx["Outlet"]] or file_outlet
            expense = str(row[idx["Expense"]]).strip()
            desk = row[idx["Deskripsi"]]
            cost_rows.append((
                no_akun,
                tgl, _norm_outlet(outlet), expense,
                str(desk).strip() if desk else None,
                float(total), source, i,
            ))
        conn.executemany(
            "INSERT INTO cost (no_akun, tanggal, outlet, expense, deskripsi,"
            " total, source_file, row_idx) VALUES (?,?,?,?,?,?,?,?)"
            " ON CONFLICT(source_file, row_idx) DO UPDATE SET"
            " no_akun=excluded.no_akun, tanggal=excluded.tanggal,"
            " outlet=excluded.outlet, expense=excluded.expense,"
            " deskripsi=excluded.deskripsi, total=excluded.total",
            cost_rows,
        )

    # --- Omset harian ---
    sheet = next((s for s in wb.sheetnames
                  if re.match(r"(?i)om(sz|s|z)et", s)), None)
    if sheet:
        ws = wb[sheet]
        try:
            hrow, idx = _header_map(ws, ["Tanggal", "Total Jasa", "Sales", "Total Omset"])
        except ValueError:
            print(f"WARNING: {source}: header omzet tidak ditemukan di sheet '{sheet}'")
        else:
            rev_rows = []
            max_idx = max(idx.values())
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if i <= hrow or len(row) <= max_idx:
                    continue
                tgl = _iso(row[idx["Tanggal"]])
                if not tgl:
                    continue
                rev_rows.append((
                    tgl, _norm_outlet(file_outlet),
                    _num(row[idx["Total Jasa"]]),
                    _num(row[idx["Sales"]]),
                    _num(row[idx["Total Omset"]]),
                    source, i,
                ))
        conn.executemany(
            "INSERT INTO revenue (tanggal, outlet, total_service, total_sales,"
            " total_omzet, source_file, row_idx) VALUES (?,?,?,?,?,?,?)"
            " ON CONFLICT(source_file, row_idx) DO UPDATE SET"
            " tanggal=excluded.tanggal, outlet=excluded.outlet,"
            " total_service=excluded.total_service,"
            " total_sales=excluded.total_sales,"
            " total_omzet=excluded.total_omzet",
            rev_rows,
        )
    # --- Summary bulanan (sumber KPI otoritatif) ---
    sum_sheet = next((s for s in wb.sheetnames
                      if s.strip().lower() == "summary"), None)
    period_m = re.match(r"(\d{2})\.", path.name)
    period_y = re.search(r"_(\d{4})\.xlsx$", path.name)
    if sum_sheet and period_m and period_y:
        period = f"{period_y.group(1)}-{period_m.group(1)}"
        ws = wb[sum_sheet]
        vals = {}
        for row in ws.iter_rows(values_only=True):
            for i, c in enumerate(row):
                if c in ("Total Service", "Total Sales", "Total Omzet",
                         "Expenses", "NETT PROFIT") and len(row) > i + 1:
                    v = row[i + 1]
                    if isinstance(v, (int, float)):
                        vals[c] = float(v)
        conn.execute(
            "INSERT INTO summary (source_file, outlet, period, total_service,"
            " total_sales, total_omzet, expenses, nett_profit)"
            " VALUES (?,?,?,?,?,?,?,?)"
            " ON CONFLICT(source_file) DO UPDATE SET"
            " outlet=excluded.outlet, period=excluded.period,"
            " total_service=excluded.total_service,"
            " total_sales=excluded.total_sales,"
            " total_omzet=excluded.total_omzet, expenses=excluded.expenses,"
            " nett_profit=excluded.nett_profit",
            (source, _norm_outlet(file_outlet), period,
             vals.get("Total Service", 0), vals.get("Total Sales", 0),
             vals.get("Total Omzet", 0), vals.get("Expenses", 0),
             vals.get("NETT PROFIT", 0)),
        )

    wb.close()
    return len(cost_rows), len(rev_rows)


def run_ingest():
    init_db()
    files = sorted(LAPORAN_DIR.rglob("*.xlsx"))
    with connect() as conn:
        for f in files:
            n_cost, n_rev = ingest_file(f, conn)
            print(f"{f.relative_to(LAPORAN_DIR)}: cost={n_cost} revenue={n_rev}")
    print("Selesai.")


_PENJUALAN_COLS = (
    "order_no", "tanggal", "jam", "order_source", "serve_by",
    "customer_type", "customer_name", "item_group", "item_name",
    "qty", "currency", "price", "discount_percent", "discount_amount",
    "amount", "service_charge", "tax_amount", "cost_perunit",
    "total_cost", "profit", "comission", "payment_type", "order_status",
    "posting_time", "source_file", "row_idx",
)

_PENJUALAN_HDR_KEYS = (
    "order no", "order date", "order time", "order source", "serve by",
    "customer type", "customer name", "item group", "item name",
    "qty", "currency", "price", "discount percent", "discount amount",
    "amount", "service charge", "tax_amount", "cost perunit",
    "total cost", "profit", "comission", "payment type", "order status",
    "posting time",
)

_PENJUALAN_NUMERIC = frozenset({
    "qty", "price", "discount_percent", "discount_amount", "amount",
    "service_charge", "tax_amount", "cost_perunit", "total_cost",
    "profit", "comission",
})

# schema col name -> header key (header key is _PENJUALAN_HDR_KEYS entry)
_PENJUALAN_NUMERIC_HDR = {
    "qty": "qty",
    "price": "price",
    "discount_percent": "discount percent",
    "discount_amount": "discount amount",
    "amount": "amount",
    "service_charge": "service charge",
    "tax_amount": "tax_amount",
    "cost_perunit": "cost perunit",
    "total_cost": "total cost",
    "profit": "profit",
    "comission": "comission",
}


def _str(v):
    return str(v).strip() if v is not None else None


def _date_str(v):
    """Return YYYY-MM-DD from a date/datetime or date-like string, else None."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date().isoformat()
        except ValueError:
            return None
    return None


def _time_hms(v):
    """Format order time as HH:MM:SS string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%H:%M:%S")
    if isinstance(v, date):
        return None
    s = str(v).strip()
    return s if s else None


def ingest_penjualan_file(path: Path, conn):
    source = str(path.relative_to(PENJUALAN_DIR))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Row 0 = header
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c else "" for c in next(rows_iter)]
    idx = {name: header.index(name) for name in _PENJUALAN_HDR_KEYS}

    data_rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        order_date_v = row[idx["order date"]]
        tanggal = _date_str(order_date_v)
        if not tanggal:
            continue

        if (_str(row[idx["item group"]]) or "").lower() == "top up":
            continue

        vals = {
            "order_no": _str(row[idx["order no"]]),
            "tanggal": tanggal,
            "jam": _time_hms(row[idx["order time"]]),
            "order_source": _str(row[idx["order source"]]),
            "serve_by": _str(row[idx["serve by"]]),
            "customer_type": _str(row[idx["customer type"]]),
            "customer_name": _str(row[idx["customer name"]]),
            "item_group": _str(row[idx["item group"]]),
            "item_name": _str(row[idx["item name"]]),
            "currency": _str(row[idx["currency"]]),
            "payment_type": _str(row[idx["payment type"]]),
            "order_status": _str(row[idx["order status"]]),
            "posting_time": _str(row[idx["posting time"]]),
            "source_file": source,
            "row_idx": i,
        }
        for col in _PENJUALAN_NUMERIC:
            vals[col] = _num(row[idx[_PENJUALAN_NUMERIC_HDR[col]]])

        data_rows.append(tuple(vals[c] for c in _PENJUALAN_COLS))

    placeholders = ",".join(["?"] * len(_PENJUALAN_COLS))
    col_names = ",".join(_PENJUALAN_COLS)
    non_key = [c for c in _PENJUALAN_COLS if c not in ("source_file", "row_idx")]
    update_set = ", ".join(f"{c}=excluded.{c}" for c in non_key)

    conn.executemany(
        f"INSERT INTO penjualan ({col_names}) VALUES ({placeholders})"
        f" ON CONFLICT(source_file, row_idx) DO UPDATE SET {update_set}",
        data_rows,
    )

    wb.close()
    return len(data_rows)


def run_ingest_penjualan():
    init_db()
    with connect() as conn:
        for path in sorted(PENJUALAN_DIR.glob("*.xlsx")):
            count = ingest_penjualan_file(path, conn)
            print(f"{path.name}: {count}")
    print("Selesai.")
